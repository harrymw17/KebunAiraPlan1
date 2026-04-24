"""
excel_reader.py
Membaca Jadwal_Kebun_Aira_Mei2026-Apr2027.xlsx dan menghasilkan
rencana mingguan berdasarkan tanggal saat ini.
"""

import os
from datetime import date, timedelta
from typing import Optional
import openpyxl

# Path ke file Excel (relatif dari folder KebunBot)
EXCEL_PATH = os.path.join(os.path.dirname(__file__), "..", "Jadwal_Kebun_Aira_Mei2026-Apr2027.xlsx")

# Mapping (tahun, bulan) → label kolom di Excel
MONTH_LABELS = {
    (2026, 5):  "Mei-26",
    (2026, 6):  "Jun-26",
    (2026, 7):  "Jul-26",
    (2026, 8):  "Ags-26",
    (2026, 9):  "Sep-26",
    (2026, 10): "Okt-26",
    (2026, 11): "Nov-26",
    (2026, 12): "Des-26",
    (2027, 1):  "Jan-27",
    (2027, 2):  "Feb-27",
    (2027, 3):  "Mar-27",
    (2027, 4):  "Apr-27",
}

MONTH_NAMES_ID = {
    1: "Januari", 2: "Februari", 3: "Maret", 4: "April",
    5: "Mei", 6: "Juni", 7: "Juli", 8: "Agustus",
    9: "September", 10: "Oktober", 11: "November", 12: "Desember"
}


def get_next_monday(ref_date: date = None) -> date:
    """Dapat tanggal Senin minggu depan dari tanggal referensi (default: hari ini)."""
    if ref_date is None:
        ref_date = date.today()
    days_ahead = 7 - ref_date.weekday()  # 0=Mon, 6=Sun
    if days_ahead == 7:
        days_ahead = 0
    return ref_date + timedelta(days=days_ahead)


def get_week_of_month(d: date) -> int:
    """Hitung minggu ke berapa dalam bulan (1–4+)."""
    first_day = d.replace(day=1)
    dom = d.day
    adjusted_dom = dom + first_day.weekday()
    return int((adjusted_dom - 1) / 7) + 1


def parse_tasks_from_text(text: str) -> list[str]:
    """Pecah teks kegiatan menjadi daftar tugas individual."""
    if not text or str(text).strip() in ("", "—", "-", "None"):
        return []
    text = str(text)
    # Split by common separators
    separators = ["; ", ", ", "\n", "→", "•"]
    tasks = [text]
    for sep in separators:
        new_tasks = []
        for t in tasks:
            parts = t.split(sep)
            new_tasks.extend(parts)
        tasks = new_tasks
    # Clean up
    cleaned = []
    for t in tasks:
        t = t.strip().strip("–—-•").strip()
        if t and len(t) > 5:
            cleaned.append(t)
    return cleaned


def read_ringkasan(wb, month_key: tuple) -> dict:
    """Baca sheet Ringkasan untuk bulan tertentu."""
    ws = wb["Ringkasan"]
    month_label = MONTH_LABELS.get(month_key, "")
    result = {}

    for row in ws.iter_rows(values_only=True):
        if row[0] and str(row[0]).startswith(month_label.split("-")[0]):
            # Match bulan (e.g., "Mei-26" matches "Mei-26")
            if month_label in str(row[0]):
                result = {
                    "bulan": str(row[0]) if row[0] else "",
                    "musim": str(row[1]) if row[1] else "",
                    "fokus": str(row[2]) if row[2] else "",
                    "kegiatan": str(row[3]) if row[3] else "",
                    "peringatan": str(row[4]) if row[4] else "",
                    "beban": str(row[5]) if row[5] else "",
                }
                break
    return result


def read_crop_tasks(wb, month_key: tuple) -> list[str]:
    """Baca sheet Crop_per_Batch untuk kegiatan bulan tertentu."""
    ws = wb["Crop_per_Batch"]
    month_label = MONTH_LABELS.get(month_key, "")
    tasks = []

    # Cari kolom yang sesuai dengan bulan
    header_row = None
    month_col = None
    for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
        for idx, cell in enumerate(row):
            if cell and month_label in str(cell):
                month_col = idx
                header_row = row
                break
        if month_col is not None:
            break

    if month_col is None:
        return tasks

    # Ekstrak tugas dari kolom bulan tersebut
    for row in ws.iter_rows(min_row=3, values_only=True):
        kategori = str(row[0]).strip() if row[0] else ""
        batch_info = str(row[1]).strip() if row[1] else ""
        if month_col < len(row):
            cell_val = row[month_col]
            if cell_val and str(cell_val).strip() not in ("", "—", "-", "None"):
                val = str(cell_val).strip()
                # Skip header rows dan rows kosong
                if kategori and not batch_info and len(val) < 30:
                    continue
                if val and val not in ("Awal Kemarau", "Kemarau", "Musim Hujan", "Puncak Kemarau",
                                        "Transisi / Awal Hujan", "Transisi / Akhir Hujan", "Puncak Hujan"):
                    # Buat label singkat
                    if batch_info and "Batch" in batch_info:
                        batch_short = batch_info.split("(")[0].strip().split("–")[0].strip()
                        task_str = f"[{batch_short}] {val}"
                    else:
                        task_str = val
                    if len(task_str) > 10:
                        tasks.append(task_str)

    return tasks


def read_land_tasks(wb, month_key: tuple) -> list[str]:
    """Baca sheet Land_Mgmt untuk kegiatan bulan tertentu."""
    ws = wb["Land_Mgmt"]
    month_label = MONTH_LABELS.get(month_key, "")
    tasks = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and month_label in str(row[0]):
            aktivitas = str(row[2]).strip() if len(row) > 2 and row[2] else ""
            if aktivitas and aktivitas not in ("None", "", "—"):
                tasks.append(aktivitas)

    return tasks


def read_people_notes(wb, month_key: tuple) -> list[str]:
    """Baca sheet People_Mgmt untuk catatan tenaga kerja bulan tertentu."""
    ws = wb["People_Mgmt"]
    month_label = MONTH_LABELS.get(month_key, "")
    notes = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and month_label in str(row[0]):
            cuti = str(row[4]).strip() if len(row) > 4 and row[4] else ""
            komunikasi = str(row[5]).strip() if len(row) > 5 and row[5] else ""
            if cuti and cuti not in ("None", "", "—"):
                notes.append(f"📅 Cuti/Libur: {cuti}")
            if komunikasi and komunikasi not in ("None", "", "—"):
                notes.append(f"💬 {komunikasi}")

    return notes


def get_weekly_plan(target_date: date = None) -> dict:
    """
    Hasilkan rencana mingguan berdasarkan tanggal target.
    Jika tidak ada tanggal, gunakan Senin minggu depan.

    Returns dict dengan:
    - week_start, week_end (date)
    - month_key (tuple)
    - bulan, musim, fokus
    - tasks (list of str)
    - warnings (list of str)
    - people_notes (list of str)
    - beban_kerja (str)
    - in_schedule (bool) - apakah dalam periode jadwal
    """
    if target_date is None:
        target_date = get_next_monday()

    week_start = target_date
    week_end = target_date + timedelta(days=6)
    month_key = (target_date.year, target_date.month)

    # Cek apakah dalam periode jadwal
    in_schedule = month_key in MONTH_LABELS

    if not in_schedule:
        return {
            "week_start": week_start,
            "week_end": week_end,
            "month_key": month_key,
            "in_schedule": False,
            "tasks": [],
            "warnings": [],
            "people_notes": [],
            "bulan": MONTH_NAMES_ID.get(target_date.month, str(target_date.month)),
            "musim": "-",
            "fokus": "Di luar periode jadwal Mei 2026–April 2027",
            "beban_kerja": "-",
        }

    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
    except FileNotFoundError:
        return {
            "week_start": week_start,
            "week_end": week_end,
            "month_key": month_key,
            "in_schedule": True,
            "error": f"File Excel tidak ditemukan di: {EXCEL_PATH}",
            "tasks": [],
            "warnings": [],
            "people_notes": [],
            "bulan": MONTH_NAMES_ID.get(target_date.month, ""),
            "musim": "",
            "fokus": "",
            "beban_kerja": "",
        }

    # Baca data dari masing-masing sheet
    ringkasan = read_ringkasan(wb, month_key)
    crop_tasks = read_crop_tasks(wb, month_key)
    land_tasks = read_land_tasks(wb, month_key)
    people_notes = read_people_notes(wb, month_key)

    # Gabungkan semua tugas, prioritaskan berdasarkan minggu ke berapa
    week_num = get_week_of_month(target_date)
    all_tasks = []

    # Tambahkan tugas dari Ringkasan (kegiatan kunci bulan ini)
    if ringkasan.get("kegiatan"):
        parsed = parse_tasks_from_text(ringkasan["kegiatan"])
        all_tasks.extend(parsed[:8])  # Maks 8 tugas dari ringkasan

    # Tambahkan tugas land & crop yang belum ada
    for t in land_tasks[:4]:
        if not any(t[:20] in existing for existing in all_tasks):
            all_tasks.append(t)

    # Filter berdasarkan minggu (tampilkan subset yang relevan)
    tasks_per_week = max(4, len(all_tasks) // 4)
    start_idx = (week_num - 1) * tasks_per_week
    week_tasks = all_tasks[start_idx:start_idx + tasks_per_week + 2]
    if not week_tasks:
        week_tasks = all_tasks[:6]  # Fallback: tampilkan 6 teratas

    # Peringatan
    warnings = []
    if ringkasan.get("peringatan"):
        raw_warn = ringkasan["peringatan"]
        for w in raw_warn.split("."):
            w = w.strip()
            if len(w) > 10:
                warnings.append(w)

    return {
        "week_start": week_start,
        "week_end": week_end,
        "month_key": month_key,
        "in_schedule": True,
        "bulan": ringkasan.get("bulan", MONTH_NAMES_ID.get(target_date.month, "")),
        "musim": ringkasan.get("musim", ""),
        "fokus": ringkasan.get("fokus", ""),
        "tasks": week_tasks,
        "warnings": warnings,
        "people_notes": people_notes,
        "beban_kerja": ringkasan.get("beban", ""),
        "week_num": week_num,
    }


def format_date_range(start: date, end: date) -> str:
    """Format rentang tanggal, e.g. '5–11 Mei 2026'"""
    if start.month == end.month:
        return f"{start.day}–{end.day} {MONTH_NAMES_ID[start.month]} {start.year}"
    return (f"{start.day} {MONTH_NAMES_ID[start.month]} – "
            f"{end.day} {MONTH_NAMES_ID[end.month]} {end.year}")
